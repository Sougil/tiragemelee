import os
import random
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

app = Flask(__name__, static_folder='static')
CORS(app)  # Autoriser les requêtes cross-origin


def generate_teams(player_count, team_size, match_count):
    """Tirage aléatoire des équipes pour un concours"""
    players = list(range(1, player_count + 1))  # Numéros des joueurs
    matches = []

    for _ in range(match_count):
        random.shuffle(players)
        match_teams = [players[i:i + team_size] for i in range(0, len(players), team_size)]
        if len(match_teams[-1]) < team_size:  # Cas où une équipe serait incomplète
            extras = match_teams.pop()
            for i, player in enumerate(extras):
                match_teams[i % len(match_teams)].append(player)
        matches.append(match_teams)

    return matches


def create_tournament_file(matches, team_size, match_count, player_count):
    """Créer un fichier Excel structuré avec les matchs et le classement global"""
    wb = Workbook()

    # Créer un onglet par partie
    for match_index, match in enumerate(matches, start=1):
        ws = wb.create_sheet(title=f"Partie {match_index}")
        
        # Entêtes des colonnes
        ws.append(["Équipe", *[f"Joueur {i+1}" for i in range(team_size)], 
                   "Résultat Équipe 1", "Résultat Équipe 2", 
                   "Points Équipe 1", "Points Équipe 2"])

        # Style des entêtes
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=6 + team_size):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajouter les équipes
        row_index = 2
        colors = [PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Jaune
                  PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")]  # Bleu clair

        for team_id, team in enumerate(match, start=1):
            ws.append([f"Équipe {team_id}", *team, "", "", "", ""])
            # Appliquer les couleurs
            for col in range(2, 2 + team_size):
                ws.cell(row=row_index, column=col).fill = colors[team_id - 1]
            row_index += 1

    # Onglet Résultat Global
    ws_global = wb.create_sheet(title="Résultat Global")
    ws_global.append(["Numéro du joueur", "Total des points", "Nombre de fois 13 points marqués", 
                      "Classement", "Indication d'ex-aequo"])

    for col in ws_global.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/generate-tournament', methods=['POST'])
def generate_tournament():
    try:
        # Récupérer les données JSON envoyées par le client
        data = request.json
        if not data:
            return jsonify({"error": "Requête JSON manquante ou mal formée"}), 400

        team_type = data.get('teamType')  # 2 pour doublette, 3 pour triplette
        player_count = data.get('playerCount')
        match_count = data.get('matchCount')

        # Validation des données
        if not all([team_type, player_count, match_count]):
            return jsonify({"error": "Tous les champs requis doivent être fournis"}), 400

        if player_count < 4:
            return jsonify({"error": "Le nombre de joueurs doit être au minimum de 4"}), 400

        if match_count < 1:
            return jsonify({"error": "Le nombre de matchs doit être au minimum de 1"}), 400

        # Vérifier la compatibilité entre le type d'équipe et le nombre de joueurs
        min_players = team_type * 2
        if player_count < min_players:
            return jsonify({
                "error": f"Pour une {'doublette' if team_type == 2 else 'triplette'}, il faut au moins {min_players} joueurs."
            }), 400

        # Tirage aléatoire des équipes
        try:
            matches = generate_teams(player_count, team_type, match_count)
        except Exception as e:
            return jsonify({"error": f"Erreur lors du tirage des équipes : {str(e)}"}), 500

        # Création du fichier Excel
        try:
            wb = create_tournament_file(matches, team_type, match_count, player_count)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
        except Exception as e:
            return jsonify({"error": f"Erreur lors de la création du fichier Excel : {str(e)}"}), 500

        return send_file(output, as_attachment=True, download_name='concours_petanque.xlsx')

    except Exception as e:
        return jsonify({"error": f"Erreur interne : {str(e)}"}), 500


if __name__ == '__main__':
    # Port configuré pour Render, avec un fallback à 5000 en local
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
